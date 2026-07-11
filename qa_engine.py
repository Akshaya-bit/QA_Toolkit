"""
Meticulous QA Test-Case Engine
================================
Converts a plain-text or PDF BRD into an interactive Excel test suite
using Google Gemini (google-generativeai) + openpyxl.

Run with:
    streamlit run qa_engine.py
"""

import io
import os

import streamlit as st
from google import genai
import pypdf
import openpyxl
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import (
    Alignment, Border, Font, PatternFill, Side
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


# ══════════════════════════════════════════════════════════════════════════════
#  PAGE CONFIG
# ══════════════════════════════════════════════════════════════════════════════
st.set_page_config(
    page_title="Meticulous QA Engine",
    page_icon="📋",
    layout="centered",
)


# ══════════════════════════════════════════════════════════════════════════════
#  CUSTOM CSS  (clean, professional look)
# ══════════════════════════════════════════════════════════════════════════════
st.markdown(
    """
    <style>
    /* ── General ── */
    [data-testid="stAppViewContainer"] { background: #F5F7FA; }
    [data-testid="stHeader"]           { background: transparent; }

    /* ── Hero banner ── */
    .hero {
        background: linear-gradient(135deg, #1F497D 0%, #2E75B6 100%);
        border-radius: 14px;
        padding: 28px 32px 22px;
        margin-bottom: 28px;
        color: #fff;
    }
    .hero h1  { margin: 0 0 6px; font-size: 1.7rem; font-weight: 700; }
    .hero p   { margin: 0; opacity: .82; font-size: .95rem; }

    /* ── Cards ── */
    .card {
        background: #fff;
        border-radius: 12px;
        padding: 22px 24px;
        margin-bottom: 18px;
        box-shadow: 0 2px 12px rgba(0,0,0,.07);
    }
    .card h3 { margin: 0 0 14px; font-size: 1rem; color: #1F497D; font-weight: 600; }

    /* ── Metric pills ── */
    .metric-row { display: flex; gap: 14px; flex-wrap: wrap; margin-bottom: 20px; }
    .metric-pill {
        flex: 1; min-width: 110px;
        background: #E9EDF4; border-radius: 10px;
        padding: 14px 16px; text-align: center;
    }
    .metric-pill .num { font-size: 1.6rem; font-weight: 700; color: #1F497D; }
    .metric-pill .lbl { font-size: .75rem; color: #595959; margin-top: 2px; }

    /* ── Status badge ── */
    .badge-ok  { background:#E6F4EA; color:#2D7D46; border-radius:6px; padding:4px 10px; font-size:.8rem; font-weight:600; }
    .badge-err { background:#FDECEA; color:#C5221F; border-radius:6px; padding:4px 10px; font-size:.8rem; font-weight:600; }

    /* ── Divider ── */
    hr.soft { border:none; border-top:1px solid #E8ECF0; margin:22px 0; }

    /* ── Primary button override ── */
    div.stButton > button[kind="primary"] {
        background: linear-gradient(135deg, #1F497D, #2E75B6);
        color: #fff; border: none; border-radius: 8px;
        padding: 10px 28px; font-weight: 600; font-size: 1rem;
        width: 100%;
        transition: opacity .2s;
    }
    div.stButton > button[kind="primary"]:hover { opacity: .88; }

    /* ── Download button ── */
    div.stDownloadButton > button {
        border-radius: 8px; font-weight: 600;
        width: 100%; padding: 10px;
    }
    </style>
    """,
    unsafe_allow_html=True,
)


# ══════════════════════════════════════════════════════════════════════════════
#  HERO BANNER
# ══════════════════════════════════════════════════════════════════════════════
st.markdown(
    """
    <div class="hero">
        <h1>📋 Meticulous QA Test-Case Engine</h1>
        <p>Upload your BRD → get a fully structured, interactive Excel test suite in seconds.</p>
    </div>
    """,
    unsafe_allow_html=True,
)


# ══════════════════════════════════════════════════════════════════════════════
#  SIDEBAR  (configuration)
# ══════════════════════════════════════════════════════════════════════════════
with st.sidebar:
    st.markdown("### ⚙️ Configuration")
    api_key = st.text_input(
        "Gemini API Key",
        value=os.environ.get("GEMINI_API_KEY", ""),
        type="password",
        help="Get a free key at https://aistudio.google.com",
    )
    st.markdown("---")
    model_choice = st.selectbox(
        "Gemini model",
        [
            "gemini-2.5-flash-preview-05-20",   # Best quality — recommended
            "gemini-2.5-flash-lite-preview-06-17",  # 2.5 Lite — faster, lighter
            "gemini-2.5-pro-preview-05-06",     # Most powerful (slower)
            "gemini-2.0-flash",                 # Fast + reliable free quota
            "gemini-2.0-flash-lite",            # Lightest, highest free quota
            "gemini-1.5-flash",                 # Fallback for older accounts
        ],
        index=0,
        help=(
            "2.5 Flash Preview is recommended. "
            "If you get a 404 error, try gemini-2.0-flash instead."
        ),
    )
    st.markdown("---")
    st.markdown(
        "<small>**Note:** Your API key is never stored or logged — "
        "it is used only for the current request.</small>",
        unsafe_allow_html=True,
    )


# ══════════════════════════════════════════════════════════════════════════════
#  MAIN FORM
# ══════════════════════════════════════════════════════════════════════════════
def extract_text(uploaded) -> str:
    """Return plain text from a .txt or .pdf upload."""
    name = uploaded.name.lower()
    if name.endswith(".pdf"):
        raw_bytes = uploaded.read()
        reader = pypdf.PdfReader(io.BytesIO(raw_bytes))
        pages_text = []
        for page in reader.pages:
            text = page.extract_text()
            if text and text.strip():
                pages_text.append(text.strip())
        if not pages_text:
            st.error(
                "❌ Could not extract text from this PDF. "
                "It may be a scanned image — please use a text-based PDF."
            )
            st.stop()
        return "\n\n".join(pages_text)
    else:
        return uploaded.read().decode("utf-8", errors="replace")


st.markdown('<div class="card"><h3>📄 Upload Business Requirements Document</h3>', unsafe_allow_html=True)

uploaded_file = st.file_uploader(
    "Drop your BRD here — PDF or plain text (.pdf / .txt)",
    type=["pdf", "txt"],
    label_visibility="collapsed",
)

if uploaded_file:
    file_ext = uploaded_file.name.lower().split(".")[-1].upper()
    st.markdown(
        f'<span class="badge-ok">📎 {file_ext} detected — ready to process</span>',
        unsafe_allow_html=True,
    )
    st.markdown("<br>", unsafe_allow_html=True)

    # Extract and preview without consuming the stream for the main run
    preview_text = extract_text(uploaded_file)
    uploaded_file.seek(0)  # reset so extract_text can run again below

    with st.expander("Preview BRD content", expanded=False):
        st.text(preview_text[:3000] + (" …" if len(preview_text) > 3000 else ""))

st.markdown("</div>", unsafe_allow_html=True)

generate_clicked = st.button("🚀 Generate Test Suite", type="primary")


# ══════════════════════════════════════════════════════════════════════════════
#  VALIDATION
# ══════════════════════════════════════════════════════════════════════════════
if generate_clicked:
    errors = []
    if not api_key:
        errors.append("Gemini API key is missing — add it in the sidebar.")
    if not uploaded_file:
        errors.append("No BRD file uploaded.")

    if errors:
        for e in errors:
            st.error(f"❌ {e}")
        st.stop()


    # ══════════════════════════════════════════════════════════════════════════
    #  STEP 1: CALL GEMINI
    # ══════════════════════════════════════════════════════════════════════════
    brd_content = extract_text(uploaded_file)

    PROMPT = f"""
You are a Senior QA Architect and Website Information Architect.

Step 1 — Silently map the logical, chronological user journey (sitemap) from the BRD below.
Step 2 — Build an exhaustive manual test suite following that chronological order.

STRICT RULES:
• Complete ALL tests for Page 1 before writing any test for Page 2.
• For EVERY page generate tests covering:
  1. UI Layout & Visuals
  2. Input Fields, Character Limits, & Boundary Checks
  3. Happy Paths (successful submissions / flows)
  4. Error Paths (blank inputs, invalid data, button spamming, session loss)

OUTPUT FORMAT — RAW TSV ONLY:
Output ONLY tab-separated rows. No markdown, no code fences, no headers,
no explanations, no blank lines. Start immediately with the first data row.

Column layout (tab-separated):
  TC-ID \\t Page Name \\t Test Case Description \\t Expected Result

Example row:
TC-HOME-001\\tHome Page\\tVerify hero banner loads within 2 s on 4 G\\tHero image and CTA visible within 2 s

---
BRD:
{brd_content}
"""

    with st.spinner("🤖 Gemini is analysing your BRD and generating test cases …"):
        client = genai.Client(api_key=api_key)
        raw_tsv = None
        last_error = None
        MAX_RETRIES = 4
        RETRY_DELAYS = [3, 6, 12, 20]   # seconds between each attempt

        for attempt in range(MAX_RETRIES):
            try:
                response = client.models.generate_content(
                    model=model_choice,
                    contents=PROMPT,
                )
                raw_tsv = response.text
                break   # success — exit retry loop
            except Exception as exc:
                last_error = exc
                if attempt < MAX_RETRIES - 1:
                    wait = RETRY_DELAYS[attempt]
                    st.toast(
                        f"⚠️ Gemini busy — retrying in {wait}s "
                        f"(attempt {attempt + 1}/{MAX_RETRIES})",
                        icon="⏳",
                    )
                    import time
                    time.sleep(wait)
                else:
                    st.error(
                        f"❌ Gemini API error after {MAX_RETRIES} attempts: {last_error}\n\n"
                        "💡 **Try these fixes in order:**\n"
                        "1. Switch to **gemini-2.0-flash** in the sidebar\n"
                        "2. Switch to **gemini-2.0-flash-lite** (highest free quota)\n"
                        "3. Wait a few minutes and try again\n"
                        "4. Check your API key is from a fresh project at aistudio.google.com"
                    )
                    st.stop()


    # ── Parse TSV ─────────────────────────────────────────────────────────────
    test_rows: list[tuple[str, str, str, str]] = []

    for line in raw_tsv.strip().splitlines():
        line = line.strip()
        # Skip markdown artifacts or empty lines
        if not line or line.startswith("```") or line.startswith("#"):
            continue

        delimiter = "\t" if "\t" in line else ","
        parts = [p.strip() for p in line.split(delimiter)]

        # Need at least TC-ID + page name
        if len(parts) < 2:
            continue

        tc_id    = parts[0] if parts[0] else f"TC-{len(test_rows)+1:03d}"
        page     = parts[1] if len(parts) > 1 else "Unknown"
        desc     = parts[2] if len(parts) > 2 else ""
        expected = parts[3] if len(parts) > 3 else ""

        test_rows.append((tc_id, page, desc, expected))

    if not test_rows:
        st.error("❌ Gemini returned no parseable test rows. Try a different model or check your BRD.")
        st.stop()


    # ══════════════════════════════════════════════════════════════════════════
    #  STEP 2: BUILD EXCEL WORKBOOK
    # ══════════════════════════════════════════════════════════════════════════
    with st.spinner("📊 Building Excel workbook …"):

        # ── Shared styles ─────────────────────────────────────────────────────
        NAVY        = "1F497D"
        NAVY_LIGHT  = "E9EDF4"
        CARD_GREY   = "F2F2F2"
        WHITE       = "FFFFFF"
        GREEN       = "2D7D46"
        RED_SOFT    = "C5221F"

        def font(size=10, bold=False, color=None, name="Segoe UI"):
            return Font(name=name, size=size, bold=bold,
                        color=color or "000000")

        def fill(hex_color):
            return PatternFill(start_color=hex_color, end_color=hex_color,
                               fill_type="solid")

        def thin_border():
            s = Side(style="thin", color="BFBFBF")
            return Border(left=s, right=s, top=s, bottom=s)

        def centre(wrap=False):
            return Alignment(horizontal="center", vertical="center",
                             wrap_text=wrap)

        def left_mid(wrap=False):
            return Alignment(horizontal="left", vertical="center",
                             wrap_text=wrap)

        # ── Workbook ──────────────────────────────────────────────────────────
        wb         = openpyxl.Workbook()
        ws_tc      = wb.active
        ws_tc.title = "Test Cases"
        ws_dash    = wb.create_sheet(title="QA Dashboard", index=0)
        ws_dash.sheet_view.showGridLines = False

        last_data_row = len(test_rows) + 1   # row index of last data row

        # ════════════════════════════════════════════════════════════════════
        #  TEST CASES SHEET
        # ════════════════════════════════════════════════════════════════════

        # Freeze header row
        ws_tc.freeze_panes = "A2"

        # ── Column widths ──
        col_widths = [18, 22, 60, 55, 16, 14]
        for i, w in enumerate(col_widths, 1):
            ws_tc.column_dimensions[get_column_letter(i)].width = w

        ws_tc.row_dimensions[1].height = 32

        # ── Header row ────────────────────────────────────────────────────
        headers = [
            "Test Case ID", "Page Name", "Test Case Description",
            "Expected Result", "Test Status", "Test Result",
        ]
        for col_idx, h in enumerate(headers, 1):
            cell = ws_tc.cell(row=1, column=col_idx, value=h)
            cell.font      = font(size=11, bold=True, color=WHITE)
            cell.fill      = fill(NAVY)
            cell.alignment = centre(wrap=True)
            cell.border    = thin_border()

        # ── Data rows ─────────────────────────────────────────────────────
        ALT_ROW = "F0F4FA"

        for row_offset, (tc_id, page, desc, expected) in enumerate(test_rows, 2):
            row_fill = fill(ALT_ROW) if row_offset % 2 == 0 else fill(WHITE)

            values = [tc_id, page, desc, expected, "Pending", "NA"]
            for col_idx, val in enumerate(values, 1):
                cell = ws_tc.cell(row=row_offset, column=col_idx, value=val)
                cell.font      = font(size=10)
                cell.fill      = row_fill
                cell.border    = thin_border()
                cell.alignment = left_mid(wrap=(col_idx in (3, 4)))

            ws_tc.row_dimensions[row_offset].height = 42

        # ── Dropdowns for Status & Result ─────────────────────────────────
        dv_status = DataValidation(
            type="list", formula1='"Tested,Pending,Blocked,NA"',
            allow_blank=True, showErrorMessage=False,
        )
        dv_result = DataValidation(
            type="list", formula1='"Pass,Fail,Skip,NA"',
            allow_blank=True, showErrorMessage=False,
        )
        ws_tc.add_data_validation(dv_status)
        ws_tc.add_data_validation(dv_result)
        dv_status.sqref = f"E2:E{last_data_row + 1}"
        dv_result.sqref = f"F2:F{last_data_row + 1}"

        # ── Auto-filter ───────────────────────────────────────────────────
        ws_tc.auto_filter.ref = f"A1:F{last_data_row}"

        # ════════════════════════════════════════════════════════════════════
        #  DASHBOARD SHEET
        # ════════════════════════════════════════════════════════════════════

        # Column widths for dashboard aesthetics
        for col_letter, w in [("A", 3), ("B", 18), ("C", 18),
                               ("D", 4), ("E", 18), ("F", 18),
                               ("G", 4), ("H", 18), ("I", 18)]:
            ws_dash.column_dimensions[col_letter].width = w

        # ── Title ─────────────────────────────────────────────────────────
        ws_dash.merge_cells("B2:I2")
        title_cell = ws_dash["B2"]
        title_cell.value     = "QA EXECUTION & STATUS DASHBOARD"
        title_cell.font      = Font(name="Segoe UI", size=16, bold=True,
                                    color=NAVY)
        title_cell.alignment = left_mid()
        ws_dash.row_dimensions[2].height = 38

        # ── KPI cards (row 4–5) ───────────────────────────────────────────
        DR = last_data_row   # last data row reference

        kpis = [
            ("TOTAL TEST CASES",    "B", "C",
             f"=COUNTA('Test Cases'!A2:A{DR})", "0"),
            ("% TESTED",            "E", "F",
             f'=IFERROR(COUNTIF(\'Test Cases\'!E2:E{DR},"Tested")/B5,0)',
             "0%"),
            ("PASSED",              "H", "I",
             f'=COUNTIF(\'Test Cases\'!F2:F{DR},"Pass")', "0"),
        ]

        for title, c1, c2, formula, num_fmt in kpis:
            # Label row
            ws_dash.merge_cells(f"{c1}4:{c2}4")
            lbl = ws_dash[f"{c1}4"]
            lbl.value     = title
            lbl.font      = Font(name="Segoe UI", size=9, bold=True,
                                 color="595959")
            lbl.alignment = centre()
            lbl.fill      = fill(NAVY_LIGHT)
            lbl.border    = thin_border()

            # Value row
            ws_dash.merge_cells(f"{c1}5:{c2}5")
            val_cell = ws_dash[f"{c1}5"]
            val_cell.value          = formula
            val_cell.font           = Font(name="Segoe UI", size=22,
                                          bold=True, color=NAVY)
            val_cell.alignment      = centre()
            val_cell.fill           = fill(CARD_GREY)
            val_cell.number_format  = num_fmt
            val_cell.border         = thin_border()

            # Apply border to both cells in merged ranges
            for r in (4, 5):
                for c in (c1, c2):
                    ws_dash[f"{c}{r}"].border = thin_border()

        ws_dash.row_dimensions[4].height = 24
        ws_dash.row_dimensions[5].height = 52

        # ── Hidden data for charts (rows 24–27) ───────────────────────────
        #  Execution status
        ws_dash["B24"] = "Status";        ws_dash["C24"] = "Count"
        ws_dash["B25"] = "Tested"
        ws_dash["C25"] = (f'=COUNTIF(\'Test Cases\'!E2:E{DR},"Tested")')
        ws_dash["B26"] = "Pending"
        ws_dash["C26"] = (f'=COUNTIF(\'Test Cases\'!E2:E{DR},"Pending")')
        ws_dash["B27"] = "Blocked"
        ws_dash["C27"] = (f'=COUNTIF(\'Test Cases\'!E2:E{DR},"Blocked")')

        #  Test results
        ws_dash["E24"] = "Result";        ws_dash["F24"] = "Count"
        ws_dash["E25"] = "Pass"
        ws_dash["F25"] = (f'=COUNTIF(\'Test Cases\'!F2:F{DR},"Pass")')
        ws_dash["E26"] = "Fail"
        ws_dash["F26"] = (f'=COUNTIF(\'Test Cases\'!F2:F{DR},"Fail")')
        ws_dash["E27"] = "Skip"
        ws_dash["F27"] = (f'=COUNTIF(\'Test Cases\'!F2:F{DR},"Skip")')

        # Style the hidden rows so they don't confuse users
        for row in range(24, 28):
            ws_dash.row_dimensions[row].height = 16
            for col in ("B", "C", "E", "F"):
                cell = ws_dash[f"{col}{row}"]
                cell.font = Font(name="Segoe UI", size=9, color="AAAAAA")

        # ── Chart 1: Execution status ──────────────────────────────────────
        chart1 = BarChart()
        chart1.type         = "col"
        chart1.style        = 10
        chart1.title        = "Test Completion Status"
        chart1.y_axis.title = "Cases"
        chart1.legend       = None
        chart1.width        = 13
        chart1.height       = 7

        chart1.add_data(
            Reference(ws_dash, min_col=3, min_row=24, max_row=27),
            titles_from_data=True,
        )
        chart1.set_categories(
            Reference(ws_dash, min_col=2, min_row=25, max_row=27)
        )
        ws_dash.add_chart(chart1, "B7")

        # ── Chart 2: Test results ─────────────────────────────────────────
        chart2 = BarChart()
        chart2.type         = "col"
        chart2.style        = 13
        chart2.title        = "Test Quality Breakdown"
        chart2.y_axis.title = "Cases"
        chart2.legend       = None
        chart2.width        = 13
        chart2.height       = 7

        chart2.add_data(
            Reference(ws_dash, min_col=6, min_row=24, max_row=27),
            titles_from_data=True,
        )
        chart2.set_categories(
            Reference(ws_dash, min_col=5, min_row=25, max_row=27)
        )
        ws_dash.add_chart(chart2, "F7")

        # ── Save to in-memory buffer ───────────────────────────────────────
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)


    # ══════════════════════════════════════════════════════════════════════════
    #  SUCCESS UI
    # ══════════════════════════════════════════════════════════════════════════

    # Compute quick stats for the UI
    pages_found   = len({r[1] for r in test_rows})
    total_cases   = len(test_rows)

    st.success("✅ Test suite generated successfully!")

    st.markdown(
        f"""
        <div class="metric-row">
            <div class="metric-pill">
                <div class="num">{total_cases}</div>
                <div class="lbl">Test Cases</div>
            </div>
            <div class="metric-pill">
                <div class="num">{pages_found}</div>
                <div class="lbl">Pages Covered</div>
            </div>
            <div class="metric-pill">
                <div class="num">2</div>
                <div class="lbl">Sheets</div>
            </div>
            <div class="metric-pill">
                <div class="num">100%</div>
                <div class="lbl">Pending</div>
            </div>
        </div>
        """,
        unsafe_allow_html=True,
    )

    # Preview table (first 10 rows)
    with st.expander("👀 Preview first 10 test cases", expanded=True):
        import pandas as pd
        preview_df = pd.DataFrame(
            test_rows[:10],
            columns=["TC ID", "Page", "Description", "Expected Result"],
        )
        st.dataframe(preview_df, use_container_width=True, hide_index=True)

    st.download_button(
        label="📥 Download interactive_qa_suite.xlsx",
        data=excel_buffer,
        file_name="interactive_qa_suite.xlsx",
        mime=(
            "application/vnd.openxmlformats-officedocument"
            ".spreadsheetml.sheet"
        ),
        use_container_width=True,
    )

    st.markdown("<hr class='soft'>", unsafe_allow_html=True)
    st.markdown(
        "<small>💡 <b>Tip:</b> Open the <b>QA Dashboard</b> sheet first — "
        "use the dropdowns in <b>Test Status</b> and <b>Test Result</b> "
        "columns to track execution live. Charts update automatically.</small>",
        unsafe_allow_html=True,
    )
